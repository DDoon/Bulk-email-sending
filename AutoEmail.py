from Account import MY_ID, MY_PW
from email.message import EmailMessage
from smtplib import SMTP_SSL
from pathlib import Path
from openpyxl import load_workbook


def send_mail(받는사람, 제목, 본문, 첨부파일, 첨부파일2, 첨부파일3=False):

    # 템플릿 생성
    msg = EmailMessage()

    # 보내는 사람 / 받는 사람 / 제목 입력
    msg["From"] = MY_ID
    msg["To"] = 받는사람.split()
    msg["Subject"] = 제목

    # 본문 구성
    msg.set_content(본문)
    
    # 파일 첨부
    if 첨부파일:
        파일명 = Path(첨부파일).name
        with open(첨부파일, "rb") as f:
            msg.add_attachment(f.read(), maintype="application", subtype="octet-stream", filename=파일명)
            
    if 첨부파일:
        파일명 = Path(첨부파일2).name
        with open(첨부파일2, "rb") as f:
            msg.add_attachment(f.read(), maintype="application", subtype="octet-stream", filename=파일명)

            
    if 첨부파일:
        파일명 = Path(첨부파일3).name
        with open(첨부파일3, "rb") as f:
            msg.add_attachment(f.read(), maintype="application", subtype="octet-stream", filename=파일명)

        
    with SMTP_SSL("smtp.naver.com", 465) as smtp:
        smtp.login(MY_ID, MY_PW)
        smtp.send_message(msg)
    
    # 완료 메시지
    print(받는사람, "clear", sep="\t")

        
# 엑셀파일에서 가져온 정보를 활용해 함수 반복 실행
wb = load_workbook('list39.xlsx', data_only=True)
ws = wb.active

for 행 in ws.iter_rows(min_row=2):
    받는사람 = 행[0].value
    제목 = 행[1].value
    본문 = 행[2].value
    첨부파일 = 행[3].value
    첨부파일2 = 행[4].value
    첨부파일3 = 행[5].value
    
    
    send_mail(받는사람, 제목, 본문, 첨부파일, 첨부파일2, 첨부파일3)