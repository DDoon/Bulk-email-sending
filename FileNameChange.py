# 파일 이름 변경(text.pdf -> text-홍길동.pdf)
# 이름만 넣어줄 때 유용하나 그냥 엑셀에 있는 기능을 사용하는게 훨 나음
import re
from openpyxl import load_workbook
import os

# load the Excel workbook
# 엑셀 파일 선택
workbook = load_workbook('file.xlsx')

# select the worksheet you want to use
# 엑셀 시트 선택
sheet = workbook['Sheet1']

# get a list of all PDF files in the directory
files = [f for f in os.listdir('파일 경로') if f.endswith('.pdf')]

# loop through the rows in the worksheet
# 파일 array정렬 1,2,3,4... 안해주면 1,10,11 순으로 감
files.sort(key=lambda x: int(re.search(r'\d+', x).group()))

for i, row in enumerate(sheet.iter_rows(values_only=True)):
    # get the value of the first column in the row
    # 엑셀 row 데이터 선택
    name = row[4]

    # get the old file name and construct the new file name
    old_file_name = files[i]
    new_file_name = old_file_name[:-4] + name + '.pdf' # 뒤에다가 예전파일이름 + 선택한 엑셀 데이터 + .pdf
    # use the os module to rename the file
    os.rename(old_file_name, new_file_name) # 바꿔준다

# save the workbook
workbook.save('file.xlsx')